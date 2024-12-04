from flask import Flask, request, jsonify, send_file
import pandas as pd
import os
from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

app = Flask(__name__)

# Функция для разделения столбца "Line ID"
def split_line_id(line_id):
    if '&erid=' in line_id:
        erid = line_id.split('&erid=')[1]
        line_id = line_id.split('&erid=')[0]
    else:
        erid = None

    if line_id.isdigit():
        return pd.Series([line_id, line_id, line_id, erid])

    if line_id.startswith('_adx_proxy_'):
        placement_and_site = "adx_proxy"
        creative_id = line_id.split('_')[-1]
        return pd.Series([placement_and_site, placement_and_site, creative_id, erid])

    parts = line_id.split('_')
    if len(parts) == 1:
        return pd.Series([parts[0], parts[0], None, erid])
    elif len(parts) == 2:
        return pd.Series([parts[0], parts[0], parts[1], erid])
    elif len(parts) == 3:
        return pd.Series([parts[0], parts[1], parts[2], erid])

    return pd.Series([None, None, None, None])

# Основная обработка файла
def process_file(file_path):
    df = pd.read_csv(file_path)

    # Разделяем столбец "Line ID"
    new_columns = df['Line ID'].apply(split_line_id)
    new_columns.columns = ['Site ID', 'Placement ID', 'Creative ID', 'Erid']

    # Удаляем уже существующие столбцы "Site ID", "Placement ID", "Creative ID", "Erid" (если они есть)
    for col in ['Site ID', 'Placement ID', 'Creative ID', 'Erid']:
        if col in df.columns:
            df = df.drop(columns=[col])

    # Добавляем новые столбцы
    df[['Site ID', 'Placement ID', 'Creative ID', 'Erid']] = new_columns

    # Удаляем дублирующиеся столбцы (на всякий случай)
    df = df.loc[:, ~df.columns.duplicated()]

    # Переставляем столбцы после "Line ID"
    line_id_index = df.columns.get_loc('Line ID')
    ordered_columns = list(df.columns[:line_id_index + 1]) + ['Site ID', 'Placement ID', 'Creative ID', 'Erid']
    remaining_columns = [col for col in df.columns if col not in ordered_columns]
    df = df[ordered_columns + remaining_columns]

    # Добавляем новые столбцы
    df['Imps GIVT, %'] = ((df['Impressions (GIVT)'] / df['Impressions (Gross)']) * 100).round(2)
    df['Imps GIVT, %'] = df['Imps GIVT, %'].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else "nan%")

    df['Clicks GIVT, %'] = ((df['Clicks (GIVT)'] / df['Clicks (Gross)']) * 100).round(2)
    df['Clicks GIVT, %'] = df['Clicks GIVT, %'].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else "nan%")

    df['Viewability, %'] = ((df['Viewable Impression'] / df['Recordable Impression']) * 100).round(2)
    df['Viewability, %'] = df['Viewability, %'].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else "nan%")

    # Переставляем новые столбцы
    if 'Impressions (Gross)' in df.columns:
        df.insert(df.columns.get_loc('Impressions (Gross)') + 1, 'Imps GIVT, %', df.pop('Imps GIVT, %'))
    if 'Clicks (Gross)' in df.columns:
        df.insert(df.columns.get_loc('Clicks (Gross)') + 1, 'Clicks GIVT, %', df.pop('Clicks GIVT, %'))
    if 'Viewable Impression' in df.columns:
        df.insert(df.columns.get_loc('Viewable Impression') + 1, 'Viewability, %', df.pop('Viewability, %'))

    # Сохраняем результат в Excel
    output_filename = f"processed_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = os.path.join('output', output_filename)
    os.makedirs('output', exist_ok=True)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')

    # Подсветка ячеек
    workbook = load_workbook(output_path)
    worksheet = workbook.active

    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    if 'Placement Type' in df.columns:
        for row in range(2, len(df) + 2):
            placement_type = worksheet.cell(row=row, column=df.columns.get_loc('Placement Type') + 1).value

            for column_name in ['Imps GIVT, %', 'Clicks GIVT, %', 'Viewability, %']:
                column_index = df.columns.get_loc(column_name) + 1
                cell = worksheet.cell(row=row, column=column_index)
                value = cell.value

                if value == "nan%":
                    continue

                percentage = float(value.strip('%'))

                if placement_type == "In-Stream Video":
                    if column_name == 'Imps GIVT, %':
                        if percentage <= 1:
                            cell.fill = green_fill
                        elif 1 < percentage <= 3:
                            cell.fill = yellow_fill
                        else:
                            cell.fill = red_fill
                    elif column_name == 'Clicks GIVT, %':
                        if percentage <= 10:
                            cell.fill = green_fill
                        elif 10 < percentage <= 20:
                            cell.fill = yellow_fill
                        else:
                            cell.fill = red_fill
                    elif column_name == 'Viewability, %':
                        if percentage >= 85:
                            cell.fill = green_fill
                        elif 65 <= percentage < 85:
                            cell.fill = yellow_fill
                        else:
                            cell.fill = red_fill

                elif placement_type in ["In-Banner"]:
                    if column_name == 'Imps GIVT, %':
                        if percentage < 1:
                            cell.fill = green_fill
                        elif 1 <= percentage <= 3:
                            cell.fill = yellow_fill
                        else:
                            cell.fill = red_fill
                    elif column_name == 'Clicks GIVT, %':
                        if percentage < 10:
                            cell.fill = green_fill
                        elif 10 <= percentage <= 20:
                            cell.fill = yellow_fill
                        else:
                            cell.fill = red_fill
                    elif column_name == 'Viewability, %':
                        if percentage > 70:
                            cell.fill = green_fill
                        elif 55 <= percentage <= 70:
                            cell.fill = yellow_fill
                        else:
                            cell.fill = red_fill

    workbook.save(output_path)
    return output_path

# Загружаем страницу для загрузки файла
@app.route("/", methods=["GET"])
def index():
    return '''
    <!doctype html>
    <title>Upload CSV File</title>
    <h1>Upload CSV File</h1>
    <form method="POST" action="/process" enctype="multipart/form-data">
        <input type="file" name="file">
        <input type="submit" value="Process">
    </form>
    '''

# Обрабатываем файл и отправляем результат
@app.route("/process", methods=["POST"])
def process():
    file = request.files['file']
    if file and file.filename.endswith('.csv'):
        input_path = os.path.join('uploads', file.filename)
        os.makedirs('uploads', exist_ok=True)
        file.save(input_path)

        output_path = process_file(input_path)
        return send_file(output_path, as_attachment=True)
    else:
        return jsonify({"error": "Invalid file format. Please upload a CSV file."}), 400

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)